'''
Created on March 6, 2019

@author: youwu
'''

import sys
from pip._vendor.distlib.compat import raw_input

import pandas as pd

from openpyxl import load_workbook 
from openpyxl.styles import colors 
from openpyxl.styles import Font, Color 

import unicodedata

class Brand:
    def __init__(self, bName='', cterm='', eterm='', oterm1='', oterm2=''):
        self.name = bName
        self.ChineseTerm = cterm
        self.EnglishTerm = eterm
        self.otherTerm1 = oterm1
        self.otherTerm2 = oterm2
    

def matchBrand( nameStr, brandInstant ):
    if brandInstant.ChineseTerm:
        if nameStr.find(brandInstant.ChineseTerm) != -1:
            return True
    if brandInstant.EnglishTerm:
        if nameStr.find(brandInstant.EnglishTerm)  != -1:
            return True
    if  brandInstant.otherTerm1:
        if nameStr.find(brandInstant.otherTerm1)  != -1:
            return True
    if brandInstant.otherTerm2:
        if nameStr.find(brandInstant.otherTerm2)  != -1:
            return True
            
    return False


def findBrand( nameStr, brandList):
    for b in brandList:
        if matchBrand(nameStr, b):
            return b
    
    b2 = Brand();
    
    return b2;


if __name__ == '__main__':
    
    wb = load_workbook('t:/L2 DLS APAC Instructions- JD Data.xlsx');
    print( wb.get_sheet_names());
    
    sheet1 = wb.get_sheet_by_name('Sheet1');
    sheet2 = wb.get_sheet_by_name('Brand Reference Sheet')
    
    blist=[];
    for row in range(2, sheet2.max_row+1):
        pname = sheet2.cell(row, 1).value;
        cname = sheet2.cell(row, 2).value;
        ename = sheet2.cell(row, 3).value;
        oname1 = sheet2.cell(row, 4).value;
        oname2 = sheet2.cell(row, 5).value;
        b = Brand(pname,cname,ename,oname1,oname2);
        blist.append(b);
    
    for row in range(2, sheet1.max_row+1):
        prodName = sheet1.cell(row, 4).value;
        brandName = sheet1.cell(row, 8).value;
        

        mb = findBrand(prodName,blist);
        msg = "Row #" + str(row) + " Original Brand:" + brandName + " Matched Brand:" + mb.name;
        print(msg)
        if not mb.name:
            sheet1.cell(row, 8).font = Font(color=colors.RED);
        elif brandName != mb.name:
            sheet1.cell(row, 8).value = mb.name;
            sheet1.cell(row, 8).font = Font(color=colors.GREEN);
        else:
            sheet1.cell(row, 8).font = Font(color=colors.BLUE);

   
    wb.save('t:/L2 DLS APAC Instructions- JD Data-updated.xlsx');
    pass
